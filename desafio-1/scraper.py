import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import time

#Variáveis globais
BASE_URL = "https://books.toscrape.com"
CATEGORY_URL = f"{BASE_URL}/catalogue/category/books/mystery_3/index.html"
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", os.path.dirname(os.path.abspath(__file__)))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "books.xlsx")

#Dicionário para converter notas de estrelas em números
STAR_RATING_MAP = {"One": 1, "Two": 2, "Three": 3, "Four": 4, "Five": 5}

#Função para obter a página
def get_soup(url: str) -> BeautifulSoup:
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    return BeautifulSoup(response.text, "lxml")

#Função para extrair a nota de estrelas
def extract_star_rating(article: BeautifulSoup) -> int:
    star_tag = article.find("p", class_="star-rating")
    if star_tag:
        for cls in star_tag.get("class", []):
            if cls in STAR_RATING_MAP:
                return STAR_RATING_MAP[cls]
    return 0

#Função para extrair detalhes do livro
def extract_book_detail(detail_url: str) -> dict:
    try:
        soup = get_soup(detail_url)

        description = ""
        desc_header = soup.find("div", id="product_description")
        if desc_header:
            desc_p = desc_header.find_next_sibling("p")
            if desc_p:
                description = desc_p.get_text(strip=True)

        upc = ""
        table = soup.find("table", class_="table-striped")
        if table:
            for row in table.find_all("tr"):
                th = row.find("th")
                td = row.find("td")
                if th and td and th.get_text(strip=True) == "UPC":
                    upc = td.get_text(strip=True)
                    break

        full_title = ""
        product_main = soup.find("div", class_="product_main")
        if product_main:
            h1 = product_main.find("h1")
            if h1:
                full_title = h1.get_text(strip=True)

        availability = ""
        avail_tag = soup.find("p", class_="instock availability")
        if avail_tag:
            availability = avail_tag.get_text(strip=True)

        return {
            "description": description,
            "upc": upc,
            "full_title": full_title,
            "availability_detail": availability,
        }

    except Exception as error:
        print(f"  ⚠ Erro ao acessar detalhe: {detail_url} — {error}")
        return {"description": "", "upc": "", "full_title": "", "availability_detail": ""}

#Função para raspar a página da categoria
def scrape_category_page(url: str) -> tuple[list[dict], str | None]:
    soup = get_soup(url)
    books = []

    for article in soup.find_all("article", class_="product_pod"):
        title_tag = article.find("h3").find("a")
        title = title_tag.get("title", title_tag.get_text(strip=True))

        detail_relative = title_tag.get("href", "")
        detail_url = f"{BASE_URL}/catalogue/{detail_relative.replace('../../../', '').replace('../../', '')}"

        price_text = article.find("p", class_="price_color").get_text(strip=True)
        price = float(price_text.replace("£", "").replace("Â", "").strip())

        rating = extract_star_rating(article)

        availability_tag = article.find("p", class_="instock availability")
        availability = availability_tag.get_text(strip=True) if availability_tag else "Unknown"

        books.append({
            "title": title,
            "price": price,
            "rating": rating,
            "availability": availability,
            "detail_url": detail_url,
        })

    next_url = None
    next_button = soup.find("li", class_="next")
    if next_button:
        next_link = next_button.find("a")
        if next_link:
            current_base = url.rsplit("/", 1)[0]
            next_url = f"{current_base}/{next_link.get('href', '')}"

    return books, next_url

#Função principal para raspar todos os livros da categoria Mystery
def scrape_all_mystery_books() -> list[dict]:
    all_books = []
    current_url = CATEGORY_URL
    page_num = 1

    print("=" * 60)
    print("🔍 INICIANDO SCRAPING — Categoria: Mystery")
    print("=" * 60)

    while current_url:
        print(f"\n📄 Página {page_num}: {current_url}")
        books, next_url = scrape_category_page(current_url)
        print(f"   Livros encontrados: {len(books)}")

        for i, book in enumerate(books, 1):
            print(f"   📖 [{i}/{len(books)}] {book['title'][:50]}...")
            details = extract_book_detail(book["detail_url"])

            if details["full_title"]:
                book["title"] = details["full_title"]

            book["description"] = details["description"]
            book["upc"] = details["upc"]

            if details["availability_detail"]:
                book["availability"] = details["availability_detail"]

            time.sleep(0.3)

        all_books.extend(books)
        current_url = next_url
        page_num += 1

    return all_books

#Função para salvar os dados no Excel
def save_to_excel(books: list[dict], filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mystery Books"

    headers = ["Título", "Preço (£)", "Nota (⭐)", "Disponibilidade", "UPC", "Descrição"]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(vertical="top", wrap_text=True)
    zebra_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    for row_idx, book in enumerate(books, 2):
        values = [
            book["title"],
            book["price"],
            book["rating"],
            book["availability"],
            book.get("upc", ""),
            book.get("description", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

            if row_idx % 2 == 0:
                cell.fill = zebra_fill

    column_widths = {1: 50, 2: 12, 3: 10, 4: 25, 5: 20, 6: 80}
    for col, width in column_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    print(f"\n💾 Arquivo salvo: {filepath}")

# Função para imprimir o resumo
def print_summary(books: list[dict]) -> None:
    total = len(books)
    avg_price = sum(b["price"] for b in books) / total if total > 0 else 0
    in_stock = sum(1 for b in books if "in stock" in b["availability"].lower())
    max_price = max(b["price"] for b in books) if books else 0
    min_price = min(b["price"] for b in books) if books else 0
    avg_rating = sum(b["rating"] for b in books) / total if total > 0 else 0

    print("\n" + "=" * 60)
    print("📊 RESUMO DA EXTRAÇÃO")
    print("=" * 60)
    print(f"  📚 Total de livros:             {total}")
    print(f"  💰 Preço médio:                 £{avg_price:.2f}")
    print(f"  📦 Disponíveis em estoque:      {in_stock}")
    print(f"  📈 Preço mais alto:             £{max_price:.2f}")
    print(f"  📉 Preço mais baixo:            £{min_price:.2f}")
    print(f"  ⭐ Média de estrelas:           {avg_rating:.1f}/5")
    print("=" * 60)

# Função principal
def main():
    start_time = time.time()

    books = scrape_all_mystery_books()

    if not books:
        print("❌ Nenhum livro encontrado. Verifique a conexão ou URL.")
        return

    save_to_excel(books, OUTPUT_FILE)
    print_summary(books)

    elapsed = time.time() - start_time
    print(f"\n⏱  Tempo de execução: {elapsed:.2f} segundos")
    print("✅ Scraping finalizado com sucesso!")


if __name__ == "__main__":
    main()